# Generate demo spreadsheet with employees, departments, and RBAC groups
import random
import openpyxl
from openpyxl.styles import Font, PatternFill

# Configuration
NUM_EMPLOYEES = 50
NUM_RBAC_GROUPS = 100

# Sample data
FIRST_NAMES = [
    'James', 'Mary', 'John', 'Patricia', 'Robert', 'Jennifer', 'Michael', 'Linda',
    'William', 'Barbara', 'David', 'Elizabeth', 'Richard', 'Susan', 'Joseph', 'Jessica',
    'Thomas', 'Sarah', 'Charles', 'Karen', 'Christopher', 'Nancy', 'Daniel', 'Lisa',
    'Matthew', 'Betty', 'Anthony', 'Margaret', 'Mark', 'Sandra', 'Donald', 'Ashley',
    'Steven', 'Kimberly', 'Paul', 'Emily', 'Andrew', 'Donna', 'Joshua', 'Michelle',
    'Kenneth', 'Carol', 'Kevin', 'Amanda', 'Brian', 'Dorothy', 'George', 'Melissa',
    'Edward', 'Deborah'
]

LAST_NAMES = [
    'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
    'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez', 'Wilson', 'Anderson',
    'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin', 'Lee', 'Perez', 'Thompson',
    'White', 'Harris', 'Sanchez', 'Clark', 'Ramirez', 'Lewis', 'Robinson', 'Walker',
    'Young', 'Allen', 'King', 'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores',
    'Green', 'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
    'Carter', 'Roberts'
]

DEPARTMENTS = [
    'Engineering', 'Product', 'Sales', 'Marketing', 'Finance', 
    'HR', 'Operations', 'Customer Success', 'Legal', 'IT'
]

# Generate RBAC groups
rbac_groups = [f"RBAC_{i:03d}" for i in range(1, NUM_RBAC_GROUPS + 1)]

# Create workbook
wb = openpyxl.Workbook()

# Sheet 1: Employees
ws_emp = wb.active
ws_emp.title = "Employees"

# Headers
headers_emp = ['Employee ID', 'First Name', 'Last Name', 'Full Name', 'Department', 'Email']
ws_emp.append(headers_emp)

# Style headers
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws_emp[1]:
    cell.fill = header_fill
    cell.font = header_font

# Generate employees
employees = []
for i in range(1, NUM_EMPLOYEES + 1):
    first_name = random.choice(FIRST_NAMES)
    last_name = random.choice(LAST_NAMES)
    full_name = f"{first_name} {last_name}"
    department = random.choice(DEPARTMENTS)
    email = f"{first_name.lower()}.{last_name.lower()}@company.com"
    
    emp_data = {
        'id': f"EMP{i:03d}",
        'first_name': first_name,
        'last_name': last_name,
        'full_name': full_name,
        'department': department,
        'email': email
    }
    employees.append(emp_data)
    
    ws_emp.append([emp_data['id'], first_name, last_name, full_name, department, email])

# Auto-size columns
for column in ws_emp.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws_emp.column_dimensions[column_letter].width = max_length + 2

# Sheet 2: RBAC Memberships
ws_rbac = wb.create_sheet("RBAC_Memberships")

# Headers
headers_rbac = ['Employee ID', 'Employee Name', 'RBAC Group']
ws_rbac.append(headers_rbac)

# Style headers
for cell in ws_rbac[1]:
    cell.fill = header_fill
    cell.font = header_font

# Assign random RBAC groups to each employee (mean=40, stddev=12)
rbac_memberships = []
for emp in employees:
    num_groups = int(random.gauss(40, 12))
    num_groups = max(1, min(num_groups, NUM_RBAC_GROUPS))  # Clamp between 1 and 100
    assigned_groups = random.sample(rbac_groups, num_groups)
    
    for group in assigned_groups:
        ws_rbac.append([emp['id'], emp['full_name'], group])
        rbac_memberships.append({
            'employee_id': emp['id'],
            'employee_name': emp['full_name'],
            'rbac_group': group
        })

# Auto-size columns
for column in ws_rbac.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws_rbac.column_dimensions[column_letter].width = max_length + 2

# Sheet 3: RBAC Groups Reference
ws_groups = wb.create_sheet("RBAC_Groups")

# Headers
headers_groups = ['RBAC Group', 'Description']
ws_groups.append(headers_groups)

# Style headers
for cell in ws_groups[1]:
    cell.fill = header_fill
    cell.font = header_font

# Add all RBAC groups
for group in rbac_groups:
    ws_groups.append([group, f"Access group {group}"])

# Auto-size columns
for column in ws_groups.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws_groups.column_dimensions[column_letter].width = max_length + 2

# Save workbook
filename = "employee_rbac_demo.xlsx"
wb.save(filename)

print(f"✓ Spreadsheet created: {filename}")
print(f"  - {NUM_EMPLOYEES} employees")
print(f"  - {NUM_RBAC_GROUPS} RBAC groups")
print(f"  - {len(rbac_memberships)} total memberships")
print(f"  - {len(DEPARTMENTS)} departments")