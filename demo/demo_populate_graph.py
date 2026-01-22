# Load knowledge graph from Excel spreadsheet
import requests
import openpyxl

# Configuration
GRAPH_ID = "YOUR_GRAPH_ID_HERE"
API_KEY = "YOUR_API_KEY_HERE"
SPREADSHEET_FILE = "employee_rbac_demo.xlsx"

url = f"https://prodaus.api.airia.ai/v1/ProjectGraph/{GRAPH_ID}/cypher"
headers = {
    "X-API-Key": API_KEY,
    "Content-Type": "application/json"
}

def run_query(query, description=""):
    response = requests.post(url, headers=headers, json={"query": query, "readonly": False})
    status = "✓" if response.status_code == 200 else "✗"
    print(f"{status} {description}")
    if response.status_code != 200:
        print(f"  Error: {response.text}")
    return response

# Load spreadsheet
print(f"Loading spreadsheet: {SPREADSHEET_FILE}")
wb = openpyxl.load_workbook(SPREADSHEET_FILE)

# Clear existing graph
print("\nClearing existing graph data...")
run_query("MATCH (n) DETACH DELETE n", "Clear graph")

# Load employees and departments
print("\nCreating employees and departments...")
ws_emp = wb["Employees"]
employees_data = []

for row in ws_emp.iter_rows(min_row=2, values_only=True):
    emp_id, first_name, last_name, full_name, department, email = row
    employees_data.append({
        'id': emp_id,
        'first_name': first_name,
        'last_name': last_name,
        'full_name': full_name,
        'department': department,
        'email': email
    })

# Create departments
departments = list(set([emp['department'] for emp in employees_data]))
dept_list = str(departments).replace("'", '"')
query = f"""
UNWIND {dept_list} AS dept
MERGE (d:Department {{name: dept}})
"""
run_query(query, f"Created {len(departments)} departments")

# Create employees one by one (more reliable)
for emp in employees_data:
    query = f"""
    MERGE (e:Employee {{
        id: '{emp['id']}',
        firstName: '{emp['first_name']}',
        lastName: '{emp['last_name']}',
        fullName: '{emp['full_name']}',
        email: '{emp['email']}'
    }})
    WITH e
    MATCH (d:Department {{name: '{emp['department']}'}})
    MERGE (e)-[:WORKS_IN]->(d)
    """
    run_query(query, f"Created employee {emp['id']}")

print(f"✓ Created {len(employees_data)} employees")

# Load RBAC groups
print("\nCreating RBAC groups...")
ws_groups = wb["RBAC_Groups"]
rbac_groups = []

for row in ws_groups.iter_rows(min_row=2, values_only=True):
    group_name, description = row
    rbac_groups.append({'name': group_name, 'description': description})

for grp in rbac_groups:
    query = f"""
    MERGE (g:RBACGroup {{name: '{grp['name']}'}})
    SET g.description = '{grp['description']}'
    """
    run_query(query, f"Created RBAC group {grp['name']}")

print(f"✓ Created {len(rbac_groups)} RBAC groups")

# Load RBAC memberships
print("\nCreating RBAC memberships...")
ws_rbac = wb["RBAC_Memberships"]
memberships = []

for row in ws_rbac.iter_rows(min_row=2, values_only=True):
    emp_id, emp_name, rbac_group = row
    memberships.append({'emp_id': emp_id, 'group': rbac_group})

# Create memberships
count = 0
for mem in memberships:
    query = f"""
    MATCH (e:Employee {{id: '{mem['emp_id']}'}})
    MATCH (g:RBACGroup {{name: '{mem['group']}'}})
    MERGE (e)-[:MEMBER_OF]->(g)
    """
    run_query(query, f"Created membership {count+1}/{len(memberships)}")
    count += 1

print(f"\n✓ Graph loaded successfully!")
print(f"  - {len(employees_data)} employees")
print(f"  - {len(departments)} departments")
print(f"  - {len(rbac_groups)} RBAC groups")
print(f"  - {len(memberships)} memberships")